"""Renderers. One report, four formats, and the same coverage in all of them.

### The rule every renderer here obeys

**Coverage renders before the numbers, and an empty section renders its
`empty_note` rather than nothing.** A CSV with a header and no rows, or a PDF
table with no lines under it, is exactly the artifact this product must not
produce: it looks like a clean month and it is indistinguishable from a month
nobody watched.

So each renderer starts with the window, the timezone, whether the period is
complete, and every gap — and only then the tables.

### Availability of a format is asked, not assumed

`reportlab` and `openpyxl` are declared under the `reports` extra rather than in
the base dependency set, the same way `onnxruntime` sits under `inference`.
A deployment that installs neither still exports CSV and JSON, and asking for
PDF gets a named capability error rather than an import traceback at request
time — the pattern `pos.not_configured` already established.

CSV and JSON are stdlib and are always available, so there is never a deployment
with no working export at all.
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from datetime import datetime
from typing import Any

from app.errors import CapabilityNotConfiguredError
from app.reporting.model import ExportFormat, ReportData, Section


@dataclass(frozen=True, slots=True)
class Rendered:
    """Bytes plus the two things an HTTP response needs to hand them over."""

    content: bytes
    media_type: str
    filename: str


def format_available(fmt: ExportFormat) -> tuple[bool, str]:
    """Whether this deployment can produce that format, and why not if it cannot."""
    if fmt in (ExportFormat.JSON, ExportFormat.CSV):
        return True, ""
    module = {ExportFormat.XLSX: "openpyxl", ExportFormat.PDF: "reportlab"}[fmt]
    try:
        __import__(module)
    except ImportError:
        return False, (
            f"{fmt.value.upper()} export needs the '{module}' package, which is "
            "declared under the 'reports' extra and is not installed in this "
            "deployment. CSV and JSON are always available."
        )
    return True, ""


def _slug(text: str) -> str:
    kept = [c.lower() if c.isalnum() else "-" for c in text]
    out = "".join(kept)
    while "--" in out:
        out = out.replace("--", "-")
    return out.strip("-")[:60] or "report"


def _stamp(report: ReportData) -> str:
    moment = report.generated_at or datetime.now()
    return moment.strftime("%Y%m%d-%H%M%S")


def _coverage_lines(report: ReportData) -> list[tuple[str, str]]:
    """The coverage block, as label/value pairs every renderer can lay out."""
    coverage = report.coverage
    lines = [
        ("Period", f"{coverage.since.isoformat()} — {coverage.until.isoformat()}"),
        (
            "Timezone",
            coverage.timezone
            if coverage.timezone_resolved
            else f"{coverage.timezone} (UNRESOLVED — boundaries computed in UTC)",
        ),
        ("Granularity", coverage.granularity.label),
        (
            "Coverage",
            "Complete — every source was read and the period has ended"
            if coverage.complete
            else "INCOMPLETE — see the gaps below before comparing these figures",
        ),
    ]
    if coverage.basis:
        lines.append(("Computed from", coverage.basis))
    for source in coverage.sources:
        lines.append(
            (
                f"Source · {source.source}",
                f"{source.rows} rows"
                + (" (truncated)" if source.truncated else "")
                if source.available
                else f"NOT AVAILABLE — {source.reason}",
            )
        )
    for gap in coverage.gaps:
        lines.append((f"Gap · {gap.kind}", gap.detail))
    return lines


# ── JSON ─────────────────────────────────────────────────────────────────────


def render_json(report: ReportData) -> Rendered:
    payload = json.dumps(report.as_dict(), indent=2, ensure_ascii=False)
    return Rendered(
        content=payload.encode("utf-8"),
        media_type="application/json",
        filename=f"{_slug(report.title)}-{_stamp(report)}.json",
    )


# ── CSV ──────────────────────────────────────────────────────────────────────


def render_csv(report: ReportData) -> Rendered:
    """One file, coverage first, sections after.

    A spreadsheet a person opens should not need a second file to tell them the
    period was incomplete, so the coverage block is rows in the same sheet
    rather than a sidecar nobody keeps.
    """
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)

    writer.writerow([report.title])
    writer.writerow([report.subtitle])
    writer.writerow([])
    writer.writerow(["COVERAGE"])
    for label, value in _coverage_lines(report):
        writer.writerow([label, value])

    if report.capability_state:
        writer.writerow([])
        writer.writerow(["MODULE STATE", report.capability_state])
        writer.writerow(["REASON", report.capability_reason])

    for section in report.sections:
        writer.writerow([])
        writer.writerow([section.title.upper()])
        if section.note:
            writer.writerow(["Note", section.note])
        writer.writerow([column.header for column in section.columns])
        if section.is_empty:
            # Never a bare header. A header with nothing under it is the shape
            # this whole module exists to avoid producing.
            writer.writerow([section.empty_note])
            continue
        for row in section.rows:
            writer.writerow([row.get(column.key, "") for column in section.columns])

    # BOM so Excel opens UTF-8 correctly on a double-click, which is how a
    # restaurant manager will actually open this.
    return Rendered(
        content=buffer.getvalue().encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        filename=f"{_slug(report.title)}-{_stamp(report)}.csv",
    )


# ── Excel ────────────────────────────────────────────────────────────────────


def render_xlsx(report: ReportData) -> Rendered:
    """A workbook: coverage on its own sheet, then one sheet per section."""
    available, reason = format_available(ExportFormat.XLSX)
    if not available:
        raise CapabilityNotConfiguredError(reason, details={"format": "xlsx"})

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    book = Workbook()
    heading = Font(bold=True, size=13)
    label = Font(bold=True)
    warn = Font(bold=True, color="9C1F17")
    header_fill = PatternFill("solid", fgColor="EEF1F3")
    wrap = Alignment(vertical="top", wrap_text=True)

    cover = book.active
    cover.title = "Coverage"
    cover["A1"] = report.title
    cover["A1"].font = heading
    cover["A2"] = report.subtitle
    cover["A2"].alignment = wrap

    row = 4
    for text, value in _coverage_lines(report):
        cover.cell(row=row, column=1, value=text).font = label
        cell = cover.cell(row=row, column=2, value=value)
        cell.alignment = wrap
        # The one thing a reader must not miss, made visually different rather
        # than merely present.
        if not report.coverage.complete and text in ("Coverage",):
            cell.font = warn
        if text.startswith("Gap ·") or "NOT AVAILABLE" in str(value):
            cell.font = warn
        row += 1

    if report.capability_state:
        row += 1
        cover.cell(row=row, column=1, value="Module state").font = label
        cover.cell(row=row, column=2, value=report.capability_state).font = warn
        row += 1
        cover.cell(row=row, column=1, value="Reason").font = label
        cover.cell(row=row, column=2, value=report.capability_reason).alignment = wrap

    cover.column_dimensions["A"].width = 26
    cover.column_dimensions["B"].width = 96

    used: set[str] = {"Coverage"}
    for section in report.sections:
        sheet = book.create_sheet(_sheet_name(section, used))
        sheet["A1"] = section.title
        sheet["A1"].font = heading
        cursor = 2
        if section.note:
            sheet.cell(row=cursor, column=1, value=section.note).alignment = wrap
            cursor += 1
        cursor += 1

        if section.is_empty:
            sheet.cell(row=cursor, column=1, value=section.empty_note).alignment = wrap
            sheet.column_dimensions["A"].width = 110
            continue

        for index, column in enumerate(section.columns, start=1):
            cell = sheet.cell(row=cursor, column=index, value=column.header)
            cell.font = label
            cell.fill = header_fill
        cursor += 1

        for data_row in section.rows:
            for index, column in enumerate(section.columns, start=1):
                sheet.cell(row=cursor, column=index, value=data_row.get(column.key, ""))
            cursor += 1

        for index, column in enumerate(section.columns, start=1):
            widest = max(
                [len(column.header)]
                + [len(str(r.get(column.key, ""))) for r in section.rows]
            )
            sheet.column_dimensions[get_column_letter(index)].width = min(60, widest + 4)

        # A real table, so Excel gives the reader filtering and banding for free.
        sheet.auto_filter.ref = (
            f"A{cursor - len(section.rows) - 1}:"
            f"{get_column_letter(len(section.columns))}{cursor - 1}"
        )

    stream = io.BytesIO()
    book.save(stream)
    return Rendered(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{_slug(report.title)}-{_stamp(report)}.xlsx",
    )


def _sheet_name(section: Section, used: set[str]) -> str:
    """Excel sheet names: 31 characters, no `[]:*?/\\`, and unique."""
    cleaned = "".join(c for c in section.title if c not in "[]:*?/\\")[:28].strip() or "Section"
    candidate = cleaned
    suffix = 2
    while candidate in used:
        candidate = f"{cleaned[:26]} {suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


# ── PDF ──────────────────────────────────────────────────────────────────────


def render_pdf(report: ReportData) -> Rendered:
    """A document somebody can sign, with the coverage on the first page."""
    available, reason = format_available(ExportFormat.PDF)
    if not available:
        raise CapabilityNotConfiguredError(reason, details={"format": "pdf"})

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        KeepTogether,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    ink = colors.HexColor("#171c20")
    muted = colors.HexColor("#4a555d")
    line = colors.HexColor("#c9cfd4")
    wash = colors.HexColor("#eef1f3")
    alert = colors.HexColor("#9c1f17")

    base = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=base["Title"], fontSize=20, leading=24,
        alignment=TA_LEFT, textColor=ink, spaceAfter=2,
    )
    sub_style = ParagraphStyle(
        "ReportSub", parent=base["Normal"], fontSize=9.5, leading=13, textColor=muted
    )
    h2 = ParagraphStyle(
        "ReportH2", parent=base["Heading2"], fontSize=12.5, leading=15,
        textColor=ink, spaceBefore=12, spaceAfter=4,
    )
    body = ParagraphStyle("ReportBody", parent=base["Normal"], fontSize=8.5, leading=11.5, textColor=muted)
    cell = ParagraphStyle("ReportCell", parent=base["Normal"], fontSize=8, leading=10, textColor=ink)
    # A right-aligned variant for numeric columns. The table's own ALIGN is not
    # enough: a `Paragraph` carries its own alignment and wins over the cell's,
    # so the `numeric` flag would be declared and have no visible effect.
    numeric_cell = ParagraphStyle("ReportNum", parent=cell, alignment=TA_RIGHT)
    alert_style = ParagraphStyle("ReportAlert", parent=body, textColor=alert)

    stream = io.BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title=report.title,
        author="UnityWorks Vision AI",
    )

    flow: list[Any] = [
        Paragraph(report.title, title_style),
        Paragraph(report.subtitle, sub_style),
        Spacer(1, 8),
    ]

    # ── coverage, first, always ──────────────────────────────────────────────
    coverage = report.coverage
    banner = (
        "Coverage complete — every source was read and the period has ended."
        if coverage.complete
        else "COVERAGE INCOMPLETE — read the notes below before comparing these figures."
    )
    flow.append(Paragraph(banner, body if coverage.complete else alert_style))
    flow.append(Spacer(1, 6))

    rows = [[Paragraph(f"<b>{k}</b>", cell), Paragraph(str(v), cell)] for k, v in _coverage_lines(report)]
    table = Table(rows, colWidths=[42 * mm, 132 * mm], hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LINEBELOW", (0, 0), (-1, -2), 0.25, line),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    flow.append(table)

    if report.capability_state:
        flow.append(Spacer(1, 10))
        flow.append(
            Paragraph(
                f"<b>This module is {report.capability_state.replace('_', ' ')}.</b> "
                f"{report.capability_reason}",
                alert_style,
            )
        )

    # ── sections ─────────────────────────────────────────────────────────────
    for section in report.sections:
        block: list[Any] = [Paragraph(section.title, h2)]
        if section.note:
            block.append(Paragraph(section.note, body))
            block.append(Spacer(1, 4))

        if section.is_empty:
            # The empty note, never a bare header row.
            block.append(Paragraph(f"<i>{section.empty_note}</i>", body))
            flow.append(KeepTogether(block))
            continue

        header = [
            Paragraph(f"<b>{c.header}</b>", numeric_cell if c.numeric else cell)
            for c in section.columns
        ]
        data = [header] + [
            [
                Paragraph(str(row.get(c.key, "")), numeric_cell if c.numeric else cell)
                for c in section.columns
            ]
            for row in section.rows
        ]
        width = (A4[0] - 36 * mm) / max(1, len(section.columns))
        section_table = Table(data, colWidths=[width] * len(section.columns), repeatRows=1)
        section_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), wash),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.25, line),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
                + [
                    # Both: the table aligns the cell box, the paragraph style
                    # aligns the text inside it. Either alone leaves the digits
                    # where they started.
                    ("ALIGN", (index, 0), (index, -1), "RIGHT")
                    for index, column in enumerate(section.columns)
                    if column.numeric
                ]
            )
        )
        block.append(section_table)
        # A short section is kept whole so its heading never sits alone at the
        # foot of a page; a long one has to be allowed to break, and its header
        # row repeats instead (`repeatRows=1`).
        if len(section.rows) <= 12:
            flow.append(KeepTogether(block))
        else:
            flow.extend(block)

    flow.append(Spacer(1, 12))
    flow.append(
        Paragraph(
            f"Generated {(report.generated_at or datetime.now()).isoformat()}. "
            "This retrieval has been recorded in the audit trail.",
            body,
        )
    )

    doc.build(flow)

    return Rendered(
        content=stream.getvalue(),
        media_type="application/pdf",
        filename=f"{_slug(report.title)}-{_stamp(report)}.pdf",
    )


RENDERERS = {
    ExportFormat.JSON: render_json,
    ExportFormat.CSV: render_csv,
    ExportFormat.XLSX: render_xlsx,
    ExportFormat.PDF: render_pdf,
}


def render(report: ReportData, fmt: ExportFormat) -> Rendered:
    """Synchronous and CPU-bound. Callers run it in a worker thread."""
    return RENDERERS[fmt](report)


__all__ = [
    "RENDERERS",
    "Rendered",
    "format_available",
    "render",
    "render_csv",
    "render_json",
    "render_pdf",
    "render_xlsx",
]
